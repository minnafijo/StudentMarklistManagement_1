# import openpyxl and tkinter modules
import datetime

from openpyxl import *
from tkinter import *
from tkcalendar import Calendar, DateEntry

import os

dirname = os.path.dirname(__file__)
path = dirname + '\StudentDetails.xlsx'
#path = os.path.join(dirname, '\StudentMarklistManagement\StudentDetails.xlsx')


#path = 'C:\\Users\\Minna\\Desktop\\ReDi_Python\\StudentMarklistManagement\\StudentDetails.xlsx'

global stud
global root
global root1
global sec_screen
global class_Selected
#global update_details

# globally declare wb and sheet variable
# opening the existing excel file
wb = load_workbook(path)
# create the sheet object
sheetTemp = wb.active

# Dictionary for storing the text widget references
cells = {}


def class_select():
    global class_Selected
    class_Selected = clicked.get()
    close_root1()


def name_select():
    global name_Selected
    name_Selected = nameclicked.get()
    print(name_Selected)
    read_details(name_Selected)


def view_details_screen():
    global window

    # Create an instance of tkinter frame
    window = Tk()

    # set the background colour of GUI window
    window.configure(background='light green')

    # set the title of GUI window
    window.title("View Details")

    # set the configuration of GUI window
    window.geometry("1500x800")

    df = wb[class_Selected]

    # Extract number of rows and columns
    n_rows = df.max_row
    n_cols = df.max_column

    # Extracting columns from the data and
    # creating text widget with some
    # background color
    column_names = df.columns
    i = 0
    for j, col in enumerate(column_names):
        text = Text(window, width=16, height=1, bg="#9BC2E6")
        text.grid(row=i, column=j)
        content = df.cell(row=i + 1, column=j + 1).value
        text.insert(INSERT, content)

    # adding all the other rows into the grid
    for i in range(n_rows-1):
        for j in range(n_cols):
            text = Text(window, width=16, height=1)
            text.grid(row=i + 1, column=j)
            content = df.cell(row=i+2, column=j+1).value
            text.insert(INSERT, content)
            cells[(i, j)] = text

    if update_details == TRUE:     #Save button will be visible
        # Create button, it will change label text
        submit = Button(window, text="Save", command=update_excel)
        submit.grid(row=15, column=8)

    # create a Back Button to the Menu window
    back = Button(window, text="Back to Menu", fg="Black",
                  bg="light blue", command=close_view)
    back.grid(row=15, column=1)

    window.mainloop()


def update_excel():

    df = wb[class_Selected]

    # Extract number of rows and columns
    n_rows = df.max_row
    n_cols = df.max_column
    column_names = df.columns

    """"
        When user clicks the "Save" button, modified data
        will be saved in excel file
        """
    student_updated = FALSE
    for i in range(1, n_rows):
        for j in range(1, n_cols):
            if df.cell(row=i+1, column=j+1).value != cells[(i-1, j)].get("1.0", "end-1c"):
                updated_Stud_row = i+1
                student_updated = TRUE
                df.cell(row=i+1, column=j+1).value = cells[(i-1, j)].get("1.0", "end-1c")
        if student_updated == TRUE:
            m1 = int(df.cell(row=updated_Stud_row, column=6).value)
            m2 = int(df.cell(row=updated_Stud_row, column=7).value)
            m3 = int(df.cell(row=updated_Stud_row, column=8).value)
            m4 = int(df.cell(row=updated_Stud_row, column=9).value)
            m5 = int(df.cell(row=updated_Stud_row, column=10).value)
            total = m1+m2+m3+m4+m5
            df.cell(row=updated_Stud_row, column=11).value = total
            df.cell(row=updated_Stud_row, column=12).value = calc_grade(total)
            student_updated=FALSE
                #obj=Student(df.cell(row=i+1, column=j+1).value)

    wb.save(path)


def main_screen():
    global root1
    root1 = Tk()
    # set the background colour of GUI window
    root1.configure(background='light green')

    # set the title of GUI window
    root1.title("Main Menu")

    # set the configuration of GUI window
    root1.geometry("500x300")

    # Dropdown menu options
    options = [
        "Class1",
        "Class2",
        "Class3",
        "Class4",
        "Class5",
        "Class6",
        "Class7",
        "Class8",
        "Class9",
        "Class10"
    ]

    global clicked
    # datatype of menu text
    clicked = StringVar()

    # initial menu text
    clicked.set("-")

    # create a Name label
    name = Label(root1, text="Select Class", bg="light green")
    name.pack()
    # Create Dropdown menu
    drop = OptionMenu(root1, clicked, *options)
    drop.pack()
    # Create button, it will change label text
    submit = Button(root1, text="Enter", command=class_select).pack()

    # create a Submit Button and place into the root window
    #submit = Button(root1, text="Add Details", fg="Black",
                    #bg="Red", command=add_details)
    #submit.grid(row=8, column=8)
    # start the GUI
    root1.mainloop()


def select_details_screen():
    global view_details
    view_details = Tk()
    # set the background colour of GUI window
    view_details.configure(background='light green')

    # set the title of GUI window
    view_details.title("View Details")

    # set the configuration of GUI window
    view_details.geometry("500x300")

    list_name()
    print("Selected class", class_Selected)
    print("List of students:", stud_names)


    global nameclicked
    # datatype of menu text
    nameclicked = StringVar()

    # initial menu text
    nameclicked.set("-")

    # create a Name label
    name = Label(view_details, text="Select Student", bg="light green")
    name.pack()
    # Create Dropdown menu
    drop = OptionMenu(view_details, nameclicked, *stud_names)
    drop.pack()

    # Create button, it will change label text
    submit = Button(view_details, text="Enter", command=name_select).pack()

    # create a Submit Button and place into the root window
    # submit = Button(root1, text="Add Details", fg="Black",
    # bg="Red", command=add_details)
    # submit.grid(row=8, column=8)
    # start the GUI
    view_details.mainloop()


def second_Screen():
    global sec_screen
    sec_screen = Tk()
    # set the background colour of GUI window
    sec_screen.configure(background='light green')

    # set the title of GUI window
    sec_screen.title("Marklist Management")

    # set the configuration of GUI window
    sec_screen.geometry("600x400")

    # create a Form label
    heading1 = Label(sec_screen, text="Student Marklist", bg="light green")
    heading1.grid(row=0, column=3)

    # create a Form label
    #heading2 = Label(sec_screen, text="Mark List", bg="light green")
    #heading2.grid(row=0, column=5)

    # create a Submit Button and place into the root window
    adddetails = Button(sec_screen, text="Add Details", fg="Black",
                    bg="Yellow", command = add_details)
    adddetails.grid(row=3, column=3)

    viewdetails = Button(sec_screen, text="View Details", fg="Black",
                        bg="Yellow", command=open_viewdetails)
    viewdetails.grid(row=4, column=3)

    updatedetails = Button(sec_screen, text="Update Details", fg="Black",
                         bg="Yellow", command=open_updatedetails)
    updatedetails.grid(row=5, column=3)

    #deletedetails = Button(sec_screen, text="Delete Details", fg="Black",
                           #bg="Yellow", command=select_details_screen)
    #deletedetails.grid(row=6, column=3)

    # create a Back Button to the root window
    back = Button(sec_screen, text="Back to Main", fg="Black",
                    bg="light blue", command=close_sec)
    back.grid(row=12, column=3)



def excel(sheet = "Template"):
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20
    sheet.column_dimensions['H'].width = 20
    sheet.column_dimensions['I'].width = 20
    sheet.column_dimensions['J'].width = 20
    sheet.column_dimensions['K'].width = 20
    sheet.column_dimensions['L'].width = 20

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Roll.No"
    sheet.cell(row=1, column=3).value = "Date of Birth"
    sheet.cell(row=1, column=4).value = "Class"
    sheet.cell(row=1, column=5).value = "Address"
    sheet.cell(row=1, column=6).value = "Mark1"
    sheet.cell(row=1, column=7).value = "Mark2"
    sheet.cell(row=1, column=8).value = "Mark3"
    sheet.cell(row=1, column=9).value = "Mark4"
    sheet.cell(row=1, column=10).value = "Mark5"
    sheet.cell(row=1, column=11).value = "Total"
    sheet.cell(row=1, column=12).value = "Grade"


# Driver code
def add_details():
    print("Add Details:")
    sec_screen.destroy()

    global root
    global name_field
    global rollno_field
    global dob_field
    global class_field
    global address_field
    global mark1_field
    global mark2_field
    global mark3_field
    global mark4_field
    global mark5_field
    global answer


    root = Tk()
    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("Registration form")

    # set the configuration of GUI window
    root.geometry("500x400")

    excel(sheetTemp)

    # create a text entry box
    # for typing the information
    name_field = Entry(root)
    rollno_field = Entry(root)
    dob_field = Entry(root)
    class_field = Entry(root)
    address_field = Entry(root)
    mark1_field = Entry(root)
    mark2_field = Entry(root)
    mark3_field = Entry(root)
    mark4_field = Entry(root)
    mark5_field = Entry(root)


    # create a Form label
    heading = Label(root, text="Form", bg="light green")

    # create a Name label
    name = Label(root, text="Name", bg="light green")

    # create a RollNo label
    RollNo = Label(root, text="Roll.No", bg="light green")

    # create a DoB label
    DoB = Label(root, text="Date of Birth\n(dd-mm-yyyy)", bg="light green")
    #Label(root, text="Choose a Date", background='gray61', foreground="white").pack(padx=20, pady=20)
    # Create a Calendar using DateEntry


    # create a classs label
    classs = Label(root, text="Class", bg="light green")

    # create a address label
    address = Label(root, text="Address", bg="light green")

    # create a mark1 label
    mark1 = Label(root, text="mark1", bg="light green")

    # create a mark2 label
    mark2 = Label(root, text="mark2", bg="light green")

    # create a mark3 label
    mark3 = Label(root, text="mark3", bg="light green")

    # create a mark4 label
    mark4 = Label(root, text="mark4", bg="light green")

    # create a mark5 label
    mark5 = Label(root, text="mark5", bg="light green")

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    name.grid(row=1, column=0)
    RollNo.grid(row=2, column=0)
    DoB.grid(row=3, column=0)
    classs.grid(row=4, column=0)
    address.grid(row=5, column=0)
    mark1.grid(row=6, column=0)
    mark2.grid(row=7, column=0)
    mark3.grid(row=8, column=0)
    mark4.grid(row=9, column=0)
    mark5.grid(row=10, column=0)



    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    name_field.bind("<Return>", focus1)

    # whenever the enter key is pressed
    # then call the focus2 function
    rollno_field.bind("<Return>", focus2)

    # whenever the enter key is pressed
    # then call the focus3 function
    dob_field.bind("<Return>", focus3)

    # whenever the enter key is pressed
    # then call the focus4 function
    class_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus5 function
    address_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    mark1_field.bind("<Return>", focus6)

    # whenever the enter key is pressed
    # then call the focus7 function
    mark2_field.bind("<Return>", focus7)

    # whenever the enter key is pressed
    # then call the focus8 function
    mark3_field.bind("<Return>", focus8)

    # whenever the enter key is pressed
    # then call the focus9 function
    mark4_field.bind("<Return>", focus9)

    # whenever the enter key is pressed
    # then call the focus10 function
    #mark5_field.bind("<Return>", focus10)

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    name_field.grid(row=1, column=1, ipadx="100")
    rollno_field.grid(row=2, column=1, ipadx="100")
    dob_field.grid(row=3, column=1, ipadx="100")
    class_field.grid(row=4, column=1, ipadx="100")
    address_field.grid(row=5, column=1, ipadx="100")
    mark1_field.grid(row=6, column=1, ipadx="100")
    mark2_field.grid(row=7, column=1, ipadx="100")
    mark3_field.grid(row=8, column=1, ipadx="100")
    mark4_field.grid(row=9, column=1, ipadx="100")
    mark5_field.grid(row=10, column=1, ipadx="100")

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="#9898F5", command=insert)
    submit.grid(row=11, column=1)

    # create a Back Button to the root window
    back = Button(root, text="Back to Main", fg="Black",
                    bg="#9898F5", command=close_root)
    back.grid(row=12, column=1)

    answer = Label(root, text='',bg="light green")
    answer.grid(row=13, column=1)

    # start the GUI
    root.mainloop()


def list_name():
    #load worksheet
    workSheet = wb[class_Selected]
    second_column = workSheet['A']
    # Create the list
    global stud_names
    stud_names = [cell.value for cell in second_column[1:]]
    print(stud_names)


def find_student(name: str):
    index = stud_names.index(name)
    return index+2


#to find and read student details row into an object from excel
def read_details(name: str):
    #read details
    global class_Selected
    print("read details")
    row_index = find_student(name)

    workSheet = wb[class_Selected]
    stud_row = workSheet[row_index]
    # Create the list
    stud_details = [cell.value for cell in stud_row[:]]
    print(stud_details)

#to display student list drop down menu.
def show_student():
    #display student names.
    print("show details")

#Function to calculate garde

def calc_grade(total):
    if total >= 450:
        grade = 'A'
    elif total >= 400:
        grade = 'B'
    elif total >= 350:
        grade = 'C'
    elif total >= 300:
        grade = 'D'
    else:
        grade = 'F'

    return grade

# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    name_field.delete(0, END)
    rollno_field.delete(0, END)
    dob_field.delete(0, END)
    class_field.delete(0, END)
    address_field.delete(0, END)
    mark1_field.delete(0, END)
    mark2_field.delete(0, END)
    mark3_field.delete(0, END)
    mark4_field.delete(0, END)
    mark5_field.delete(0, END)


class Student:

    # Constructor
    def __init__(self, name, rollno, DoB, classs, addr, m1:int ,m2:int ,m3:int, m4:int, m5:int):
        self.total = None
        self.grade = None
        self.name = name
        self.rollno = rollno
        self.classs = classs
        self.DoB = DoB
        self.mark1 = m1
        self.mark2 = m2
        self.mark3 = m3
        self.mark4 = m4
        self.mark5 = m5
        self.address = addr


    # Function to display student details
    def display(self):
        print("Name : ", self.name)
        print("RollNo : ", self.rollno)
        print("Marks1 : ", self.mark1)
        #print("Marks2 : ", ob.m2)
        print("\n")

    def calculate_grade(self):
        self.total = self.mark1 + self.mark2 + self.mark3 + self.mark4 + self.mark5

        if self.total >= 450:
            self.grade = 'A'
        elif self.total >= 400:
            self.grade = 'B'
        elif self.total >= 350:
            self.grade = 'C'
        elif self.total >= 300:
            self.grade = 'D'
        else:
            self.grade = 'F'


    def save(self, sheet):

        self.calculate_grade()
        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column
        sheet.cell(row=current_row + 1, column=1).value = self.name
        sheet.cell(row=current_row + 1, column=2).value = self.rollno
        sheet.cell(row=current_row + 1, column=3).value = self.DoB
        sheet.cell(row=current_row + 1, column=4).value = self.classs
        sheet.cell(row=current_row + 1, column=5).value = self.address
        sheet.cell(row=current_row + 1, column=6).value = self.mark1
        sheet.cell(row=current_row + 1, column=7).value = self.mark2
        sheet.cell(row=current_row + 1, column=8).value = self.mark3
        sheet.cell(row=current_row + 1, column=9).value = self.mark4
        sheet.cell(row=current_row + 1, column=10).value = self.mark5
        sheet.cell(row=current_row + 1, column=11).value = self.total
        sheet.cell(row=current_row + 1, column=12).value = self.grade

        # save the file
        wb.save(path)
        print("Succesfully Saved")


# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (name_field.get() == "" or
            rollno_field.get() == "" or
            dob_field.get() == "" or
            class_field.get() == "" or
            address_field.get() == "" or
            mark1_field.get() == "" or
            mark2_field.get() == "" or
            mark3_field.get() == "" or
            mark4_field.get() == "" or
            mark5_field.get() == ""):

        print("empty input")
        answer.config(text="Empty Input")

    else:

        # get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        name = name_field.get()



        addr = address_field.get()
        try:
            rollno = int(rollno_field.get())
        except ValueError:
            answer.config(text="Please enter valid rollno")

        else:

            try:
                classs = int(class_field.get())
                if classs > 10 or classs < 1:
                    raise ValueError()
            except ValueError:
                answer.config(text="Please enter valid class(1-10)")

            else:

                try:
                    m1 = int(mark1_field.get())

                    m2 = int(mark2_field.get())
                    m3 = int(mark3_field.get())
                    m4 = int(mark4_field.get())
                    m5 = int(mark5_field.get())
                    if (m1 > 100 or m1 < 0 or
                            m2 > 100 or m2 < 0 or
                            m3 > 100 or m3 < 0 or
                            m4 > 100 or m4 < 0 or
                            m5 > 100 or m5 < 0):
                        raise ValueError()

                except ValueError:
                    answer.config(text="Please enter valid marks(0-100)")

                else:
                    dob = dob_field.get()
                    try:
                        transaction_date = datetime.datetime.strptime(dob, "%d-%m-%Y")
                    except ValueError:
                        answer.config(text='Please enter date DD-MM-YYYY')
                    else:
                        #global class_Sheet
                        class_Sheet = wb["Class" + str(classs)]

                        # call excel function
                        excel(class_Sheet)

                        global stud
                        stud = Student(name, rollno, dob, classs,  addr, m1, m2, m3, m4, m5)
                        stud.save(class_Sheet)

                        # set focus on the name_field box
                        name_field.focus_set()

                        # call the clear() function
                        clear()

def display():
    stud.display()


def close_root():
    root.destroy()
    second_Screen()


def close_view():
    window.destroy()
    second_Screen()


def close_sec():
    sec_screen.destroy()
    main_screen()


def open_viewdetails():
    global update_details
    update_details = FALSE
    sec_screen.destroy()
    view_details_screen()
def open_updatedetails():
    global update_details
    update_details =TRUE
    sec_screen.destroy()
    view_details_screen()

def close_root1():
    root1.destroy()
    second_Screen()




# Function to set focus (cursor)
def focus1(event):
    # set focus on the rollno_field box
    rollno_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the dob_field box
    dob_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the classs box
    class_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the address_field box
    address_field.focus_set()


    # Function to set focus
def focus5(event):
    # set focus on the mark1_field box
    mark1_field.focus_set()


    # Function to set focus
def focus6(event):
    # set focus on the mark2_field box
    mark2_field.focus_set()


    # Function to set focus
def focus7(event):
    # set focus on the mark3_field box
    mark3_field.focus_set()

def focus8(event):
    # set focus on the mark4_field box
    mark4_field.focus_set()


    # Function to set focus
def focus9(event):
    # set focus on the mark5_field box
    mark5_field.focus_set()






